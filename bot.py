import os
import json
import base64
import logging
from io import BytesIO

import requests
import openpyxl
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes

# ─── تنظیمات ───────────────────────────────────────────────
TELEGRAM_TOKEN = os.environ["TELEGRAM_TOKEN"]
OPENAI_API_KEY = os.environ["OPENAI_API_KEY"]

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ─── استخراج اطلاعات فاکتور با OpenAI ─────────────────────
def extract_invoice_data(image_bytes: bytes) -> dict:
    image_b64 = base64.b64encode(image_bytes).decode("utf-8")

    payload = {
        "model": "gpt-4o",
        "max_tokens": 1000,
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{image_b64}"
                        }
                    },
                    {
                        "type": "text",
                        "text": """Bu faturadaki bilgileri çıkar ve SADECE aşağıdaki JSON formatında döndür, başka hiçbir şey yazma:

{
  "fatura_no": "...",
  "tarih": "...",
  "satici": "...",
  "alici": "...",
  "kalemler": [
    {"aciklama": "...", "miktar": "...", "birim_fiyat": "...", "toplam": "..."}
  ],
  "ara_toplam": "...",
  "kdv": "...",
  "genel_toplam": "...",
  "para_birimi": "..."
}

Eğer bir alan görünmüyorsa boş string yaz."""
                    }
                ]
            }
        ]
    }

    headers = {
        "Authorization": f"Bearer {OPENAI_API_KEY}",
        "Content-Type": "application/json"
    }

    response = requests.post(
        "https://api.openai.com/v1/chat/completions",
        headers=headers,
        json=payload,
        timeout=30
    )
    response.raise_for_status()

    content = response.json()["choices"][0]["message"]["content"]
    content = content.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()
    return json.loads(content)


# ─── ساخت فایل اکسل ────────────────────────────────────────
def create_excel(data: dict) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fatura"

    # استایل هدر
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E75B6")

    # اطلاعات کلی فاکتور
    info_rows = [
        ("Fatura No", data.get("fatura_no", "")),
        ("Tarih", data.get("tarih", "")),
        ("Satıcı", data.get("satici", "")),
        ("Alıcı", data.get("alici", "")),
    ]

    for i, (label, value) in enumerate(info_rows, start=1):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    # فاصله
    ws.cell(row=6, column=1, value="")

    # هدر جدول کالاها
    headers = ["Açıklama", "Miktar", "Birim Fiyat", "Toplam"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ردیف‌های کالا
    items = data.get("kalemler", [])
    for row_idx, item in enumerate(items, start=8):
        ws.cell(row=row_idx, column=1, value=item.get("aciklama", ""))
        ws.cell(row=row_idx, column=2, value=item.get("miktar", ""))
        ws.cell(row=row_idx, column=3, value=item.get("birim_fiyat", ""))
        ws.cell(row=row_idx, column=4, value=item.get("toplam", ""))

    # جمع‌ها
    summary_start = 8 + len(items) + 1
    para = data.get("para_birimi", "")
    ws.cell(row=summary_start, column=3, value="Ara Toplam").font = Font(bold=True)
    ws.cell(row=summary_start, column=4, value=f"{data.get('ara_toplam', '')} {para}")
    ws.cell(row=summary_start+1, column=3, value="KDV").font = Font(bold=True)
    ws.cell(row=summary_start+1, column=4, value=f"{data.get('kdv', '')} {para}")
    ws.cell(row=summary_start+2, column=3, value="Genel Toplam").font = Font(bold=True)
    ws.cell(row=summary_start+2, column=4, value=f"{data.get('genel_toplam', '')} {para}")

    # عرض ستون‌ها
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─── هندلرهای تلگرام ───────────────────────────────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 Merhaba! Ben Fatura Bot.\n\n"
        "📸 Fatura fotoğrafı gönderin → 📊 Excel dosyası alın!\n\n"
        "Hemen bir fatura fotoğrafı göndererek başlayın."
    )

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("⏳ Fatura işleniyor, lütfen bekleyin...")

    try:
        # دریافت عکس
        photo = update.message.photo[-1]
        file = await context.bot.get_file(photo.file_id)
        image_bytes = await file.download_as_bytearray()

        # استخراج اطلاعات
        data = extract_invoice_data(bytes(image_bytes))

        # ساخت اکسل
        excel_file = create_excel(data)

        # ارسال فایل
        await update.message.reply_document(
            document=excel_file,
            filename="fatura.xlsx",
            caption="✅ Faturanız Excel'e dönüştürüldü!"
        )

    except json.JSONDecodeError:
        await update.message.reply_text("❌ Fatura okunamadı. Lütfen daha net bir fotoğraf gönderin.")
    except Exception as e:
        logger.error(f"Hata: {e}")
        await update.message.reply_text("❌ Bir hata oluştu. Lütfen tekrar deneyin.")

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("📎 Lütfen faturayı fotoğraf olarak gönderin (dosya değil).")


# ─── اجرا ──────────────────────────────────────────────────
if __name__ == "__main__":
    app = ApplicationBuilder().token(TELEGRAM_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    print("✅ Bot çalışıyor...")
    app.run_polling()
